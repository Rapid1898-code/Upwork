import pathlib
import os
from os.path import expanduser
import shutil
from datetime import datetime
import sys
import time
import random
import subprocess
from email.mime.text import MIMEText
import smtplib
import string
from selenium import webdriver
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.webdriver.chrome.options import Options as ChromeOptions
from sys import platform
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd
from uuid import getnode as get_mac
if platform == "win32" and get_mac() != 242380407969819:
    import mysql.connector

def backup_file(folder,fn,days):
    """
    make backup-file in folder - every day one file
    :param folder: where the backup files should be stored
    :param fn: name of the file which should be backuped
    :param days: how many days back the backup should be stored
    :return: nothing
    """
    files=[]
    dates=[]
    act_dt = datetime.now ().strftime ("%Y-%m-%d")

    # check if foldername exists - if not create it
    folder_name = folder.replace("/","").replace("\\","")
    if os.path.isdir(folder_name) == False: os.mkdir("Backup")

    # check which backup-file are in the folder
    for i in os.listdir("Backup"):
        if fn in i: files.append(i)
    files.sort (reverse=True)
    # read the dates of the backups
    for i in files:
        if "#_" not in i: continue
        else: dates.append(i.split("#_")[0])
    # make backup with actual date
    if act_dt not in dates:
        fn_backup = os.getcwd() + folder + act_dt + "#_" + fn
        shutil.copy(fn,fn_backup)
        print("File backup done... ",fn_backup)
        # check if there are more backup-days then necessary and delete
        # and if yes delete the last / oldest one
        while len(files)+1 > days:
            os.remove(os.getcwd() + folder + files[-1])
            files.pop()

def vpn_switch(sek,listCountry = False):
    """
    VPN-switch for nordvpn with x seconds delay
    return-value: returns the randomly choosed country
    country-list (for best connection-speed is hardcoded)
    :param sek: waiting for vpn-switch for X seconds
    :return: returns the randomly choosed country
    """
    if listCountry == False:
        countries = ["Austria", "Belgium", "Germany", "Israel", "Italy", "Bosnia and Herzogovina",
                     "Norway", "Poland", "Portugal", "Romania", "Serbia", "Switzerland", "United Kingdom",
                     "Bulgaria", "Croatia", "Estonia", "Finland", "France", "Georgia", "Greece",
                     "Hungary", "Iceland", "Latvia", "Netherlands"]
    else:
        countries = listCountry
    rand_country = random.randrange (len (countries) - 1)
    subprocess.call (["C:/Program Files/NordVPN/NordVPN.exe", "-c", "-g", countries[rand_country]])
    print ("VPN Switch to", countries[rand_country], "...")
    for i in range (sek, 0, -1):
        sys.stdout.write (str (i) + ' ')
        sys.stdout.flush ()
        time.sleep (1)
    print ("Connected to", countries[rand_country], "...")
    return (countries[rand_country])

def send_mail(loginName, loginPW, smtpServer, smtpPort, msgTxt, headlineTxt, sender, recList):
# Send Mail
    """
    send e-mail
    :param loginName:  login name (own email adress)
    :param loginPW:  login pw (own login email pw - sometimes a created, not the real one like with gmail)
        eg. for gmail with 2-factor-authentifiation you have to generate a app-pw
        see: https://support.google.com/mail/answer/185833?hl=en
    :param smtpServer: eg. smtp.gmail.com for gmail or mail.gmx.net for gmx
    :param eg. port 587 for gmail or gmx
    :param msg_txt: text to send
    :param headline_txt:  headline to send
    :param sender: e-mail of the sender (probably the same as the loginName
    :param recList: list of recipients-emails
    :return:

    Example:
    rtt.send_mail(
        loginName = "markuspolzer73@gmail.com",
        loginPW = "*app-pw*",
        smtpServer = "smtp.gmail.com",
        smtpPort = "587",
        msgTxt = "MessageTest BlaBlaBla",
        headlineTxt = "Headline text",
        sender = "markuspolzer73@gmail.com",
        recList = ["rapid1898@gmail.com"])
    """
    # s = smtplib.SMTP ('smtp.gmail.com', 587)  # SMTP-Server and port number from the mail provider (e.g. GMail)
    # s = smtplib.SMTP ('mail.gmx.net', 587)  # SMTP-Server and port number from the mail provider (e.g. GMail)
    s = smtplib.SMTP (smtpServer, str(int(smtpPort)))
    print (s.ehlo ())  # Check if OK - Response 250 means connection is ok
    print (s.starttls ())  # Check if OK
    print (s.login (loginName, loginPW))  # Check if OK
    msg = MIMEText (msgTxt)

    # read receiver list
    recipients = recList
    msg['Subject'] = headlineTxt
    msg['From'] = sender
    msg['To'] = ", ".join (recipients)
    s.sendmail (sender, recipients, msg.as_string ())
    s.quit ()

def check_element_dict (name_v, dict_v, default_v, idx):
    """
    Check if element is in a dictionary and has a valid value
    works for dict with only single elements and also for lists
    (similar to .get-function from dict - but with false value handling)
    :param name_v: name of the key which sould be checked
    :param dict_v: dictionary which sould be checked
    :param default_v: default value which should be used when key is not in dict or value is not valid
    :param idx: if "-1" only one entry per key - if <> "-1" this is the index of the list-element from the key toc check
    :return: default value if key is not in dict or entry is not valid - otherwise nothing happens
    """
    if name_v in dict_v:
        # check if element is single element or list
        # working for dicts with only one element and also for dicts with several entries per key
        if idx == -1:
            if dict_v[name_v] not in [None, "N/A"]: return (dict_v[name_v])
            else: return(default_v)
        else:
            if dict_v[name_v][idx] not in [None,"N/A"]: return(dict_v[name_v][idx])
            else: return(default_v)
    else: return (default_v)

def printNumAbbr(value):
    """
    Make abbreviaton for numeric value in thousands (K), millions (M), billions (B) or trillions (T)
    :param value: numeric value which should be abbreviated
    :return: string value with maximum possible abbreviation
    """
    minus = False
    if value in ["N","N/A",None,""]: return("N/A")
    if str(value)[0] == "-":
        value = float(str(value).replace ("-", ""))
        minus = True
    if value > 1000000000000:
        tmp = round(value / 1000000000000,2)
        if minus: return ("-"+str(tmp)+"T")
        else: return (str(tmp)+"T")
    elif value > 1000000000:
        tmp = round (value / 1000000000, 2)
        if minus: return ("-"+str(tmp)+"B")
        else: return (str (tmp) + "B")
    elif value > 1000000:
        tmp = round (value / 1000000, 2)
        if minus: return ("-"+str(tmp)+"M")
        else: return (str (tmp) + "M")
    elif value > 1000:
        tmp = round (value / 1000, 2)
        if minus: return ("-"+str(tmp)+"K")
        else: return (str (tmp) + "K")
    else: return value

def changePerc(oldValue, newValue):
    """
    Calculates percentage change from 2 values
    with checking if both values are available / valid
    :param oldValue: old value
    :param newValue: new value
    :return: change in percentage (or "N/A" if values are not valid)
    """
    if oldValue not in ["N/A","",None] and newValue not in ["N/A","",None]:
        oldValue = float(oldValue)
        newValue = float(newValue)
        change = (newValue-oldValue) / oldValue * 100
        return(round(change,2))
    else:
        return("N/A")

def col2num(col):
    """
    Converts XLSX-column-char to XLSX column-number
    :param n: XLSX column-char
    :return: number of column
    """
    num = 0
    for c in col:
        if c in string.ascii_letters:
            num = num * 26 + (ord(c.upper()) - ord('A')) + 1
    return num

def num2col(n):
    """
    Converts numeric XLSX-column-number to XLSX column-char
    :param n: number of column to convert
    :return: XLSX column-char
    """
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string

def wait_countdown (sek):
    """
    Shows countdown in a line during waiting for X seconds
    :param sek: Number of seconds to wait
    :return: nothing
    """
    for i in range (sek, 0, -1):
        sys.stdout.write (str (i) + ' ')
        sys.stdout.flush ()
        time.sleep (1)
    print("\n")

def replace_more (inp_str, list_chars, target_char=""):
    """
    Replace several chars in a string
    :param inp_str: string which should be changed
    :param list_chars: which chars should be changed in list-form
    :param target_char: with which char the list_chars should be replaced - default is ""
    :return: changed string
    """
    for i in list_chars:
        inp_str = inp_str.replace(i,target_char)
    return inp_str

def define_driver(headless=True):
    """
    Create driver for selenium in Chrome
    Working for Windows, Linux, MacOS
    :param headless: if TRUE driver is headless (works in background) - if FALSE selenium works on screen
    :return: Created selenium driver for further working
    """
    options = ChromeOptions()
    if headless: options.add_argument ('--headless')
    options.add_experimental_option ('excludeSwitches', ['enable-logging'])
    if platform == "win32":
        driver = webdriver.Chrome (os.getcwd () + '/chromedriver.exe', options=options)
    elif platform in ["linux","darwin"]:
        driver = webdriver.Chrome (os.getcwd () + '/chromedriver', options=options)
    return(driver)

def defineDriverFF(headless=True):
    """
    TODO: is under construction now used in a running program so far...
    Create driver for selenium in Firefox
    Working for Windows, Linux, MacOS
    :param headless: if TRUE driver is headless (works in background) - if FALSE selenium works on screen
    :return: Created selenium driver for further working
    """
    options = FirefoxOptions()
    if headless: options.add_argument ('--headless')
    #options.add_experimental_option ('excludeSwitches', ['enable-logging'])
    tmp = os.getcwd ()
    if platform == "win32":
        tmp = os.path.join (tmp, 'geckodriver.exe')
        print(tmp)
    elif platform == "linux":
        tmp = os.path.join (tmp, 'geckodriver')
    print(tmp)
    driver = webdriver.Firefox (tmp, options=options)
    return(driver)

def close_popup(driver,mode,cont):
    """
    PopUp-Closer during Webscraping
    :param driver: driver for webscraping - to define see function "defineDriver"
    :param mode: different modes for element scrapping with selenium (eg. xpath, id)
    :param cont: content of the element to search (eg. with xpath or id)
    :return: TRUE if popup was closed - FALSE nothing happened
    """
    try:
        if mode == "xpath":
            driver.find_element_by_xpath (cont).click ()
            print ("PopUp Closed...")
        elif mode == "id":
            driver.find_element_by_id (cont).click ()
            print ("PopUp Closed...")
        return (True)
    except:
        return (False)

def checkFreeVersion(prg, t=3):
    """
    CheckProgram for x free tries in a program - is stored in a hidden file in the standard download folder
    :param prg: name of the program to check (don´t use the program-name - use something else like jkf5xvjvrredkmv4n
    :param t: number of free tries to use the program per day
    :return: True if free limit is not reached - otherwise return False
    """
    fn = expanduser("~/Downloads/" + prg + ".ini").replace("\\","/")
    # check if file exists - if not create it with init value
    if os.path.exists(fn) == False:
        s = ""
        s2 = ""
        for i in range (98): s += str (random.randint (0, 9))
        for i in range (14): s2 += str (random.randint (0, 9))
        initW = s + "00" + s2
        with open (fn,"a") as f: f.write(initW)
        # hide file
        p = os.popen('attrib +h ' + fn)
        p.read()
        p.close()

    with open (fn,"r+") as f:
        w = f.read()
        #print("DEBUG InitialString: ",w)
        actDateCode = int(w[18:31])
        #print("DEBUG DateCode Before: ",actDateCode)
        tries = int(w[98:100])
        #print("DEBUG Tries Before: ",tries)
        actDate = datetime.fromtimestamp(actDateCode / 1e3).date()
        #print("DEBUG Date Before: ",actDate)

        tday = datetime.today ().date ()
        if tday == actDate:
            if tries >= t: return(False)
            else:
                tries += 1
                if tries < 10:
                    tries = "0"+str(tries)
                newW = w[:98] + str(tries) + w[100:]
        # old/no date in file - initialize with actual date and 1
        else:
            tdayISO = datetime.fromisoformat (str(tday)).timestamp ()
            tdayISO = int (tdayISO) * 1000
            newW = w[:18] + str(tdayISO) + w[31:98] + "00" + w[100:]
            tries = 0

        # print(f"DEBUG New: {newW}")
        print(f"Actual {tries} from possible {t} per day...")
        # print("DEBUG Date Actual: ",tday)
        f.seek(0)
        f.write(newW)
        return(True)

def growthCalc(listElem, countElem):
    """
    Return the Growth Rate of Elements in a list
    :param listElem:    Individual elements in list (calculate growth from right value to left value)
    :param countElem:   Number of Elements for which the growth should be calculated
                        if -1 then the elements are counted automatic
    :return:            growth rate as float
    """
    initialListElem = listElem
    initialCountElem = countElem
    if countElem == -1:
        for idx in range(len(listElem)-1,-1,-1):
            if listElem[idx] in [None,"","N/A",0]:
                del listElem[idx]
        if len(listElem) == 0:
            return ("N/A")
        countElem = len(listElem) - 1
    else:
        for idx in range(len(listElem)-1,-1,-1):
            if listElem[idx] in [None,"","N/A",0]:
                del listElem[idx]
        if len(listElem) <= countElem:
            countElem = len(listElem) - 1

    listGrowth = []
    tmpGrowth = 0
    for i, e in enumerate (listElem):
        if i < len (listElem) - 1:
            try:
                tmpGrowth = (e - listElem[i + 1]) / listElem[i + 1] * 100
            except:
                print(f"Growth Calculation not possible: {e}, Content: {listElem}, ListElem: {listElem}, CountElem {countElem}, "
                      f"Initial ListElem {initialListElem}, Initial CountElem {initialCountElem}")
            listGrowth.append (tmpGrowth)
    count = 0
    sumGrowth = 0

    # print(f"DEBUG: ListGrowth: {listGrowth}")
    # print(f"DEBUG: CountElem: {countElem}")

    for i, e in enumerate (listGrowth):
        sumGrowth += e
        count += 1
        if count == countElem: break
    if count == 0 and sumGrowth == 0:
        return("N/A")
    else:
        return sumGrowth / countElem

def save_xls(worksheetName, content, filename, delWS=False):
    """
    Save a nested list in a XLSX
    Adjust the column width according to the content
    Check if the XLSX is open an ask the user to close the XLSX
    :param worksheetName: name of the worksheet to store
    :param content: nested list with the content to save
    :param filename: name of the XLSX file
    :param delWS: if TRUE deletes worksheet before saving if it allready exists
                  if FALSE add the worksheet to the existing with (nr) at the end
    :return: nothing
    """

    book = load_workbook (filename)
    if delWS and worksheetName in book.sheetnames: del book[worksheetName]
    writer = pd.ExcelWriter (filename, engine='openpyxl', options={'strings_to_numbers': True})
    writer.book = book
    pd.DataFrame (content).to_excel (writer, sheet_name=worksheetName, header=False, index=False)

    # Automatische Anpassung der Spalten nach best fit
    # Ermittlung des längsten Wertes pro Spalte
    column_widths = []
    ws = writer.sheets[worksheetName]
    for row in content:
        for i, cell in enumerate (row):
            if len (column_widths) > i:
                if len (str (cell)) > column_widths[i]:
                    column_widths[i] = len (str (cell))
            else:
                column_widths += [len (str (cell))]
    for i, column_width in enumerate (column_widths):
        ws.column_dimensions[get_column_letter (i + 1)].width = column_width + 2

    while True:
        try:
            writer.save ()
            writer.close ()
            print("Saved to",filename,"...")
            break
        except Exception as e:
            print ("Error: ", e)
            input ("Excel-File can not be opened - please save, close it and press <Enter>")

def sql_connector(hostPM, userPM, passwdPM, databasePM):
    """
    create cursor and db-connector for database
    local Maria-DB: (host="localhost", user="root", passwd="I65faue#MB7#", database="stockdb")
    local MySQL-DB: (host="localhost",user="root",passwd="I65faue#ML5#",database="stockdb")
    a2hosting.com database: (host="nl1-ss18.a2hosting.com", user="rapidtec_Rapid1898", passwd="I65faue#AG9#", database="rapidtec_stockdb")
    :param hostPM: hostname
    :param userPM: username
    :param passwdPM: password
    :param databasePM: database name
    :return: db-cursor and db-connector
    """
    mydb = mysql.connector.connect (host=hostPM, user=userPM, passwd=passwdPM, database=databasePM)
    return(mydb.cursor(),mydb)

def dcf_calc(fcf, fcf_growth, marr, netto_cash, shares_outst,moss=1):
    if fcf in [None,"N/A"] or fcf_growth in [None,"N/A"] or shares_outst in [None,"N/A"]: return(0,0,0)

    # calculate discountedcashflow fair value
    fcf = fcf
    shares_outst = shares_outst
    disc_cashflow_10y = \
        fcf * (fcf_growth + 1) / (marr + 1) \
        + fcf * (fcf_growth + 1) ** 2 / (marr + 1) ** 2 \
        + fcf * (fcf_growth + 1) ** 3 / (marr + 1) ** 3 \
        + fcf * (fcf_growth + 1) ** 4 / (marr + 1) ** 4 \
        + fcf * (fcf_growth + 1) ** 5 / (marr + 1) ** 5 \
        + fcf * (fcf_growth + 1) ** 6 / (marr + 1) ** 6 \
        + fcf * (fcf_growth + 1) ** 7 / (marr + 1) ** 7 \
        + fcf * (fcf_growth + 1) ** 8 / (marr + 1) ** 8 \
        + fcf * (fcf_growth + 1) ** 9 / (marr + 1) ** 9 \
        + fcf * (fcf_growth + 1) ** 10 / (marr + 1) ** 10

    final_terminal_value = (fcf * (fcf_growth + 1) ** 10) / \
                           (marr + 1) ** 10 * (1 + 3 / 100) / (marr - 3 / 100)

    price_dcf = round ((disc_cashflow_10y + final_terminal_value + netto_cash) / shares_outst * moss, 2)

    return (price_dcf, disc_cashflow_10y, final_terminal_value)

def isNumber(s):
    try:
        float(s)
        return True
    except ValueError:
        return False

def uppercaseDictKeys (lowerDict):
    """
    Uppercases all keys in a Dict
    :param lowerDict:   Dict which keys should be uppercased
    :return:            Dict with uppercased keys
    """
    upperDict = {}
    for k, v in lowerDict.items():
        upperDict[k.upper()] = v
    return upperDict

if __name__ == '__main__':
    GROWTH_CALC = True
    ISNUMBER = False

    if GROWTH_CALC:
        print(f"\nCheck Function <growthCalc>: ")
        print('assert growthCalc([6,None,5,4,3],-1) == 26.11111111111111')
        assert growthCalc([6,None,5,4,3],-1) == 26.11111111111111
        print ('assert growthCalc([7,6,5,4,3],-1) == 23.75')
        assert growthCalc([7,6,5,4,3],-1) == 23.75
        print('assert growthCalc([7,6,5,4,3],5) == 23.75')
        assert growthCalc([7,6,5,4,3],5) == 23.75
        print ('assert growthCalc([None,"","N/A"],-1) == "N/A"')
        assert growthCalc([None,"","N/A"],-1) == "N/A"
        print ('assert growthCalc ([], 5) == "N/A"')
        assert growthCalc ([], 5) == "N/A"
        print ('assert growthCalc ([65339, 90488, 107147, 126032], -1) == -19.441574670075422')
        assert growthCalc ([65339, 90488, 107147, 126032], -1) == -19.441574670075422
        print('assert growthCalc([0.57, 0.0, 2.92, 3.24, 1.97, 1.9],2) == -45.17799763233553')
        assert growthCalc([0.57, 0.0, 2.92, 3.24, 1.97, 1.9],2) == -45.17799763233553
        print('assert growthCalc([0.57, 0.0, 2.92, 3.24, 1.97, 1.9],5) == -5.551194915553283')
        assert growthCalc([0.57, 0.0, 2.92, 3.24, 1.97, 1.9],5) == -5.551194915553283




    if ISNUMBER:
        print (f"\nCheck Function <isNumber>: ")
        print('assert isNumber("-1.59") == True')
        assert isNumber("-1.59") == True
        print('assert isNumber("1.59") == True')
        assert isNumber("1.59") == True
        print('assert isNumber("-123321,59") == False')
        assert isNumber("-123321,59") == False

    #backup_file("/Backup/","Test.xlsx",7)
    # defineDriverFF (headless=True)

        # print(col2num("C"))
        # print(num2col(3))









